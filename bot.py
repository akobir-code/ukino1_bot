import os
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes

KINO_KODLARI = {
    "1": {"file_id": "https://t.me/toonuz/7", "caption": "Desperoni sarguzashtlari"},
    "2": {"file_id": "https://t.me/toonuz/8", "caption": "Dora va yo'qolgan shahar"},
    "3": {"file_id": "https://t.me/toonuz/9", "caption": "Zamonlar osha 1"},
    "4": {"file_id": "https://t.me/toonuz/10", "caption": "Zamonlar osha 2"},
    "5": {"file_id": "https://t.me/toonuz/11", "caption": "Zamonlar osha 3"},
    "6": {"file_id": "https://t.me/toonuz/12", "caption": "Uy hayvonlarining sirli hayoti"},
    "7": {"file_id": "https://t.me/toonuz/13", "caption": "Mikki va yangi yil ertagi"},
    "8": {"file_id": "https://t.me/toonuz/14", "caption": "Shaxzoda va bo'ri"},
    "9": {"file_id": "https://t.me/toonuz/15", "caption": "Shaxzoda va bo'ri 2"},
    "10": {"file_id": "https://t.me/toonuz/16", "caption": "Shaxzoda va bo'ri 3"},
   # "11": {"file_id": "https://t.me/toonuz/12", "caption": "Obuna bo`ling"},
   # "12": {"file_id": "https://t.me/toonuz/13", "caption": "Obuna bo`ling"},
   # "13": {"file_id": "https://t.me/toonuz/14", "caption": "Obuna bo`ling"},
   # "14": {"file_id": "https://t.me/toonuz/15", "caption": "Obuna bo`ling"},
   # "15": {"file_id": "https://t.me/toonuz/16", "caption": "Obuna bo`ling"},
   # "16": {"file_id": "https://t.me/toonuz/17", "caption": "Obuna bo`ling"},
   # "17": {"file_id": "https://t.me/toonuz/18", "caption": "Obuna bo`ling"},
   # "18": {"file_id": "https://t.me/toonuz/19", "caption": "Obuna bo`ling"},
   # "19": {"file_id": "https://t.me/toonuz/20", "caption": "Obuna bo`ling"},
   # "20": {"file_id": "https://t.me/toonuz/21", "caption": "Obuna bo`ling"},
   # "21": {"file_id": "https://t.me/toonuz/22", "caption": "Obuna bo`ling"},
   # "22": {"file_id": "https://t.me/toonuz/23", "caption": "Obuna bo`ling"},
   # "23": {"file_id": "https://t.me/toonuz/24", "caption": "Obuna bo`ling"},
   # "24": {"file_id": "https://t.me/toonuz/25", "caption": "Obuna bo`ling"},
   # "25": {"file_id": "https://t.me/toonuz/26", "caption": "Obuna boling"},
   # "26": {"file_id": "https://t.me/toonuz/27", "caption": "Obuna bo`ling"},
   # "27": {"file_id": "https://t.me/toonuz/28", "caption": "Obuna bo`ling"},
   # "28": {"file_id": "https://t.me/toonuz/29", "caption": "Obuna bo`ling"},
   # "29": {"file_id": "https://t.me/toonuz/30", "caption": "Obuna bo`ling"},
   # "30": {"file_id": "https://t.me/toonuz/31", "caption": "Obuna bo`ling"},
    
}

KANALLAR = [
    "@toonuz",
    "@multiklar_D"
]

foydalanuvchilar_ruxsati = set()
EXCEL_FILENAME = "obunachilar.xlsx"

# Excel faylga ID yozish funksiyasi
def excelga_yozish(user_id, fullname):
    if not os.path.exists(EXCEL_FILENAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Telegram ID", "Foydalanuvchi"])
        wb.save(EXCEL_FILENAME)

    wb = load_workbook(EXCEL_FILENAME)
    ws = wb.active

    # Takror yozishni oldini olish
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            wb.close()
            return

    ws.append([user_id, fullname])
    wb.save(EXCEL_FILENAME)
    wb.close()

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [[InlineKeyboardButton("🔄 Tekshirish", callback_data="check_subscription")]]
    markup = InlineKeyboardMarkup(keyboard)

    kanal_royxat = "\n".join([f"👉 {kanal}" for kanal in KANALLAR])
    await update.message.reply_text(
        f"❗️Quyidagi kanallarga obuna bo‘ling:\n\n{kanal_royxat}\n\nSo‘ngra pastdagi tugmani bosing 👇",
        reply_markup=markup
    )

async def foydalanuvchi_obuna_tekshiruvi(user_id, context):
    obuna_bolmaganlar = []
    for kanal in KANALLAR:
        try:
            chat_member = await context.bot.get_chat_member(kanal, user_id)
            if chat_member.status not in ["member", "creator", "administrator"]:
                obuna_bolmaganlar.append(kanal)
        except:
            obuna_bolmaganlar.append(kanal)
    return obuna_bolmaganlar

async def tekshirish_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    fullname = f"{query.from_user.first_name or ''} {query.from_user.last_name or ''}".strip()

    obuna_emas = await foydalanuvchi_obuna_tekshiruvi(user_id, context)

    if obuna_emas:
        royxat = "\n".join([f"❌ {kanal}" for kanal in obuna_emas])
        await query.edit_message_text(
            f"Quyidagi kanallarga hali obuna bo‘lmagansiz:\n\n{royxat}\n\nObuna bo‘lib, qayta urining /start"
        )
    else:
        foydalanuvchilar_ruxsati.add(user_id)
        excelga_yozish(user_id, fullname)
        await query.edit_message_text("✅ Obuna tasdiqlandi. Endi kino kodini yuboring.")

async def kod_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in foydalanuvchilar_ruxsati:
        await update.message.reply_text("❗️Iltimos, avval /start buyrug‘i orqali obunangizni tekshiring.")
        return

    kod = update.message.text.strip()
    kino = KINO_KODLARI.get(kod)
    if kino:
        await update.message.reply_text(f"\n{kino['file_id']}")
        # yoki fayl id bo‘lsa: await update.message.reply_video(kino['file_id'], caption=kino['caption'])
    else:
        await update.message.reply_text("❌ Bunday kod topilmadi.")

if __name__ == '__main__':
    app = ApplicationBuilder().token("7567381417:AAH23ZUSDvII1_lxoCdDEE4imppNN-Yx0Fc").build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(tekshirish_handler, pattern="check_subscription"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, kod_qabul))

    app.run_polling()
